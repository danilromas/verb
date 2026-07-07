# -*- coding: utf-8 -*-
"""Build unified engineering equipment catalog from all sources."""

import json
import re
import html
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(r"Z:\ibp")
OUT_DIR = ROOT / "catalog"
OUT_DIR.mkdir(exist_ok=True)

SECTION_MARKERS = (
    "сетевые", "мониторинг", "аксессуар", "гибрид", "однофаз", "трехфаз",
    "трёхфаз", "универсальн", "срок службы", "серия", "линейка", "группа",
    "комплект", "креплен", "кабель", "коннектор", "батаре", "ячейк",
    "модуль", "инвертор deye", "growatt", "must", "epever", "srne",
    "lifepo", "cotek", "крепления", "комплектующ", "ликвидац", "акция",
    "новинк", "раздел", "прайс", "примечан",
)


def is_section_header(name: str) -> bool:
    if not name or len(name) < 3:
        return True
    low = name.lower().strip()
    if low in ("наименование", "модель", "название", "артикул"):
        return True
    if not re.search(r"\d", name) and len(name.split()) <= 6:
        for m in SECTION_MARKERS:
            if m in low:
                return True
    if low.endswith(":") or low.endswith("!") and "growatt" not in low and "sunways" not in low:
        if not re.search(r"\d", name):
            return True
    return False


def fmt_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if abs(v - round(v)) < 0.01:
            return int(round(v))
        return round(float(v), 2)
    return v


def price_block(retail_foreign=None, retail_rub=None, dealer_foreign=None, dealer_rub=None, currency="юань"):
    block = {}
    if retail_foreign is not None or retail_rub is not None:
        block["retail"] = {
            "foreign": fmt_num(retail_foreign),
            "foreign_currency": currency,
            "rub": fmt_num(retail_rub),
        }
    if dealer_foreign is not None or dealer_rub is not None:
        block["dealer"] = {
            "foreign": fmt_num(dealer_foreign),
            "foreign_currency": currency,
            "rub": fmt_num(dealer_rub),
        }
    return block


def find_header(rows):
    for i, row in enumerate(rows[:25]):
        if not row:
            continue
        texts = [str(c).strip().lower() if c else "" for c in row]
        joined = " ".join(texts)
        if "артикул" in joined or ("цена" in joined and ("розниц" in joined or "ррц" in joined or "дилер" in joined)):
            return i, [str(c).strip() if c else "" for c in row]
    return None, None


def col_index(header, *patterns):
    for idx, h in enumerate(header):
        hl = h.lower()
        for p in patterns:
            if p in hl:
                return idx
    return None


def parse_excel():
    wb = openpyxl.load_workbook(ROOT / "Прайс_лист_SUNWAYS_29.04.26.xlsx", read_only=True, data_only=True)
    products = []
    stats = {"sheets": 0, "rows": 0, "skipped_sections": 0}

    sheet_category = {
        "Sunways FSM": ("Солнечные панели", "Sunways"),
        "One-Sun OS": ("Солнечные панели", "One-Sun"),
        "АКБ Sunways": ("Аккумуляторы", "Sunways"),
        "Инверторы Deye|Growatt": ("Инверторы", "Deye/Growatt"),
        "Инверторы": ("Инверторы", "Sunways"),
        "EPEVER": ("Инверторы и контроллеры", "EPEVER"),
        "SRNE": ("Контроллеры заряда", "SRNE"),
        "LiFePO4_LiNMC": ("Литиевые аккумуляторы", "LiFePO4"),
        "COTEK": ("Инверторы", "COTEK"),
        "Крепления СМ": ("Крепления", "Sunways"),
        "Комплектующие": ("Комплектующие", "Sunways"),
    }

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        stats["sheets"] += 1
        hi, header = find_header(rows)
        if hi is None:
            continue

        cat, manufacturer = sheet_category.get(sn, (sn, "Sunways"))
        art_i = col_index(header, "артикул")
        name_i = col_index(header, "наименование", "модель") or 1
        tech_i = col_index(header, "технолог")
        power_i = col_index(header, "pmax", "мощност", "ватт")

        retail_rub_i = col_index(header, "ррц, руб", "ррц,руб", "ррц руб")
        retail_yuan_i = col_index(header, "ррц, юань", "ррц,юань", "розница, юань", "цена розница, юань")
        retail_usd_i = col_index(header, "розница, usd", "цена розница, usd")

        dealer_yuan_i = col_index(header, "дилер, юань", "цена дилер, юань", "дилер,  юань")
        dealer_usd_i = col_index(header, "дилер, usd", "цена дилер, usd")
        dealer_rub_i = col_index(header, "дилер, в руб", "дилер в руб")

        for row in rows[hi + 1 :]:
            if not row or not any(c is not None and str(c).strip() for c in row):
                continue
            stats["rows"] += 1

            def cell(i):
                if i is None or i >= len(row):
                    return None
                v = row[i]
                if v is None or (isinstance(v, str) and not v.strip()):
                    return None
                return v

            article = str(cell(art_i)).strip() if cell(art_i) else ""
            name = str(cell(name_i)).strip() if cell(name_i) else ""
            if not name:
                continue
            if is_section_header(name):
                stats["skipped_sections"] += 1
                continue

            # One-Sun: model in col0, power in col1
            if sn == "One-Sun OS" and not article and row[0]:
                article = str(row[0]).strip()
                if name.isdigit() or (isinstance(cell(name_i), (int, float))):
                    name = f"OS-{article.split('-')[-1] if '-' in article else article}"

            model = name
            series = ""
            m = re.match(r"^(Sunways|Growatt|GoodWe|Goodwe|Deye|MUST|Epever|SR-|SRNE|LFP|SUN-)", name, re.I)
            if m:
                series = m.group(1)

            specs = []
            tech = cell(tech_i)
            if tech:
                specs.append(f"Технология: {tech}")
            power = cell(power_i)
            if power:
                specs.append(f"Пиковая мощность: {power} Вт")

            chars_in_name = re.findall(r"\(([^)]+)\)", name)
            for c in chars_in_name:
                specs.append(c)

            currency = "юань"
            retail_foreign = cell(retail_yuan_i) or cell(retail_usd_i)
            dealer_foreign = cell(dealer_yuan_i) or cell(dealer_usd_i)
            if cell(retail_usd_i) or cell(dealer_usd_i):
                currency = "USD"

            retail_rub = cell(retail_rub_i)
            dealer_rub = cell(dealer_rub_i)

            note = ""
            for c in row:
                if isinstance(c, str) and c.strip() in ("Акция!", "Ликвидация!", "Новинка!", "Хит!"):
                    note = c.strip()

            products.append({
                "id": f"price-{len(products)+1}",
                "source": "Прайс_лист_SUNWAYS_29.04.26.xlsx",
                "sheet": sn,
                "category": cat,
                "manufacturer": manufacturer,
                "series": series,
                "model": model,
                "article": article,
                "name": name,
                "description": name,
                "characteristics": "; ".join(specs) if specs else "",
                "advantages": note,
                "prices": price_block(retail_foreign, retail_rub, dealer_foreign, dealer_rub, currency),
                "images": [],
                "documentation": [],
                "country": "",
            })

    wb.close()
    return products, stats


def parse_bobrovsolar():
    path = ROOT / "Аккумуляторы 200 Ач для солнечных станций купить в Москве 👍 цены, каталоги.html"
    text = path.read_text(encoding="utf-8", errors="ignore")
    products = []
    pattern = re.compile(
        r'data-item="\{([^"]+)\}".*?itemprop="name" content="([^"]+)".*?'
        r'PROPERTY_ARTICLE_VALUE&quot;:&quot;([^&]*)&quot;.*?'
        r'itemprop="price" content="(\d+)".*?'
        r'itemprop="image"[^>]+src="([^"]+)"',
        re.DOTALL,
    )
    for m in pattern.finditer(text):
        raw_json = html.unescape(m.group(1).replace("&quot;", '"'))
        try:
            data = json.loads("{" + raw_json.split("{", 1)[-1] if not raw_json.startswith("{") else raw_json)
        except Exception:
            data = {}
        name = m.group(2)
        article = m.group(3)
        price_rub = int(m.group(4))
        img = m.group(5)
        brand = "Sunways" if "SUNWAYS" in name.upper() else name.split()[1] if len(name.split()) > 1 else ""

        products.append({
            "id": f"web-bobrov-{len(products)+1}",
            "source": "bobrovsolar.ru",
            "sheet": "Аккумуляторы 200 Ач",
            "category": "Аккумуляторы",
            "manufacturer": brand,
            "series": "",
            "model": name.replace("Аккумулятор ", ""),
            "article": article,
            "name": name,
            "description": name,
            "characteristics": "Ёмкость: 200 А·ч",
            "advantages": "",
            "prices": price_block(retail_rub=price_rub, currency="руб"),
            "images": [img],
            "documentation": [],
            "country": "",
            "url": data.get("DETAIL_PAGE_URL", ""),
        })
    return products


def get_inverter_articles():
    """Structured data from mywatt.ru inverter ranking article."""
    return [
        {
            "category_class": "Премиум-класс",
            "manufacturer": "Schneider Electric",
            "country": "Франция",
            "models": ["Conext XW+ 8548", "Conext CL20E BASE", "Conext ESSENTIAL", "Conext OPTIMUM", "XW", "SW"],
            "description": "На вопрос, какой инвертор лучше для большого загородного дома, обычно дают ответ – Schneider Electric. Эти модели недешевы, но их качество и возможности оправдывают каждый вложенный рубль. Французская компания, основанная двумя немцами, стояла у истоков промышленной революции и вышла на энергетический рынок в 1836 году. В XXI веке в состав концерна входило несколько крупнейших производителей электроники со всего мира, включая Xantrex Technology Inc. – мирового лидера в производстве электронных компонентов для систем СЭС и ВЭС. Наиболее продаваемой моделью является гибридный инвертор Schneider Electric Conext XW+ 8548 с огромным функционалом возможностей и непревзойденной надежностью. Модели XW и SW могут подмешивать, продавать и аккумулировать энергию от солнечных батарей, что не могут сделать любые сетевые инверторы.",
            "advantages": "Полное отсутствие электрохимических конденсаторов; эффективность преобразователя AC/DC – 98,0 ÷ 98,3%; три варианта модификаций – BASE, ESSENTIAL и OPTIMUM; надежный комплекс систем защиты и встроенный MPPT терминал.",
            "features": "Гибридные аккумуляторные инверторы on-grid/off-grid; масштабируемые системы для крупных домов и объектов.",
            "application": "Большие загородные дома; гибридные и резервные системы; объекты с централизованной подачей энергии параллельно с аккумуляторами.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/sexw8548.webp",
        },
        {
            "category_class": "Премиум-класс",
            "manufacturer": "Fronius",
            "country": "Австрия",
            "models": ["Fronius Primo", "Fronius Symo", "Fronius Solar.Web"],
            "description": "Бренд Fronius одноименной компании из Австрии пятый год подряд занимает уверенное первое место среди коммерческих и бытовых моделей. Лучшие инверторы этой марки отличаются уникальным сочетанием надежности, эффективности и безукоризненного обслуживания. Всем пользователям, зарегистрировавшимся на сайте Fronius Solar.Web, фирма предоставляет 10 лет гарантии на свое оборудование.",
            "advantages": "Технология активного принудительного охлаждения; система защиты от поступления постоянного тока; простая конструкция с защелкивающейся передней крышкой; мониторинг Fronius Solar.Web с Wi-Fi; бесплатное приложение Solarweb; опция Fronius Ohmpilot для водонагревателя.",
            "features": "Инновационные технологические решения; удаленный доступ и управление.",
            "application": "Бытовые и коммерческие солнечные электростанции; сетевые on-grid системы.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/fronius.webp",
        },
        {
            "category_class": "Премиум-класс",
            "manufacturer": "Outback Power",
            "country": "США",
            "models": ["OutBack FX", "OutBack VFX"],
            "description": "Outback Power Technologies — ведущий производитель профессиональных систем автономного и резервного электроснабжения. Outback FX/VFX — продвинутые сверхнадежные масштабируемые системы для автономного и резервного электроснабжения. Инверторы OutBack выпускают в двух модификациях FX и VFX.",
            "advantages": "Уникальная защита от воздействий окружающей среды; параллельное наращивание до 30 кВт (до 10 инверторов); 5-стадийное зарядное устройство; простота обслуживания на месте; минимальный шум (один вентилятор в VFX).",
            "features": "Автономные и резервные off-grid системы; микропроцессорное зарядное устройство.",
            "application": "Автономное и резервное электроснабжение; off-grid объекты; профессиональные системы.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/outbackpower.webp",
        },
        {
            "category_class": "Премиум-класс",
            "manufacturer": "SMA Solar Technologies",
            "country": "Германия",
            "models": ["Tripower Core-1", "SunnyHighpower", "SunnyIsland", "Sunny Boy AV"],
            "description": "Широко известный немецкий бренд, до 2016 года успешно боровшийся за мировое лидерство с австрийской продукцией Fronius. В среде специалистов часто называется оптимальным по соотношению цена/качество. Покупатели отмечают сетевые модификации Tripower Core-1, SunnyHighpower и SunnyIsland.",
            "advantages": "Оптимальное соотношение цена/качество; мощный чистый сигнал Wi-Fi; проверенная немецкая инженерия.",
            "features": "Сетевые инверторы; серия AV без внешнего дисплея (мониторинг через Wi-Fi).",
            "application": "Сетевые солнечные станции; коммерческие и бытовые объекты.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/sma.webp",
        },
        {
            "category_class": "Премиум-класс",
            "manufacturer": "Huawei",
            "country": "Китай",
            "models": ["Sun2000L", "Sun2000 серия 2-5 кВт"],
            "description": "Компания вышла на инверторный рынок относительно недавно. Первый выпущенный инвертор Sun2000L стал лидером по спросу, надежности и функциональности. Сегодня Huawei предлагает линейку небольших бытовых моделей весом не более 10 кг и мощностью 2-5 кВт.",
            "advantages": "Стильный дизайн; компактность; работа в автономных и гибридных системах; высокая надежность.",
            "features": "Бытовые сетевые и гибридные инверторы; легкий вес до 10 кг.",
            "application": "Домашние солнечные станции; гибридные системы 2-5 кВт.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/huawei.webp",
        },
        {
            "category_class": "Премиум-класс",
            "manufacturer": "Sungrow",
            "country": "Китай",
            "models": ["SHK5-20", "Cristal G2"],
            "description": "Шестой в списке и первый в Китае на внутреннем рынке. Бренд обязан появлением университетскому преподавателю. В премиальном классе благодаря безупречной надежности при сравнительно невысокой стоимости.",
            "advantages": "Безупречная надежность; доступная цена; портал мониторинга iSolarCloud; приложение для смартфона.",
            "features": "Гибридные модели; хорошее соотношение цена/надежность.",
            "application": "Домашние гибридные системы; объекты с удаленным мониторингом.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/sungrow.webp",
        },
        {
            "category_class": "Бюджетный класс",
            "manufacturer": "Bineos",
            "country": "Китай (Тайваньская прописка, производство Шэньчжень)",
            "models": ["Bineos EM3KF"],
            "description": "Один из лучших китайских брендов в области инверторов. При сравнительно невысокой стоимости продукция отличается безукоризненным качеством. Отдельное подразделение исследований — свыше 200 человек.",
            "advantages": "Идеальная исходящая синусоида; высокая эффективность; широкий диапазон настроек; информативность; компактность и стильный дизайн.",
            "features": "Бюджетные инверторы с чистой синусоидой.",
            "application": "Домашние автономные и гибридные системы.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/bineos_em3kf.webp",
        },
        {
            "category_class": "Бюджетный класс",
            "manufacturer": "SmartWatt",
            "country": "Россия (Delta Solar)",
            "models": ["SmartWatt ECO", "SmartWatt Hybrid", "SmartWatt Grid 5K 1P 2MPPT"],
            "description": "Российский бренд компании Delta Solar. Основные покупатели — РФ и ближнее зарубежье. Огромное разнообразие инверторов и диапазон цен от 20 до 370 тыс. руб.",
            "advantages": "ECO: 1-7,2 кВт, встроенный контроллер, обязательны АКБ; Hybrid: многофункциональные, работа без АКБ; GRID: сетевые до 60 кВт.",
            "features": "Три серии: ECO, Hybrid, GRID; широкий ценовой диапазон.",
            "application": "Гибридные, резервные и сетевые системы для дома и бизнеса.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/smartwatt.webp",
        },
        {
            "category_class": "Бюджетный класс",
            "manufacturer": "Solax Power",
            "country": "Китай (Suntellite Group)",
            "models": ["Solax Power X"],
            "description": "Новичок, быстро завоевавший популярность. Принадлежит концерну Suntellite Group. Третье место в рейтинге традиционно завоеван стоимостью — цена линейки X почти в 2 раза ниже европейских аналогов.",
            "advantages": "Низкая цена; быстрый рост популярности; качество сопоставимо с европейцами.",
            "features": "Линейка Solax Power X.",
            "application": "Домашние солнечные станции; бюджетные on-grid решения.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/solax.webp",
        },
        {
            "category_class": "Бюджетный класс",
            "manufacturer": "GoodWe",
            "country": "Китай",
            "models": ["GoodWe DNS 2019"],
            "description": "Китайский бренд с стильным дизайном и надежной работой. За 10 лет не получил объективных претензий по качеству. GoodWe DNS 2019 — самый востребованный: надежный, компактный и дешевый.",
            "advantages": "Фантастически надежная работа; стильный дизайн; отсутствие претензий по качеству за 10 лет.",
            "features": "Компактные сетевые инверторы без излишних функций.",
            "application": "Бюджетные домашние солнечные станции.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/goodwe.webp",
        },
        {
            "category_class": "Бюджетный класс",
            "manufacturer": "Delta",
            "country": "Китай (Delta Group)",
            "models": ["Delta Home Series", "Delta RPI (устаревшая)"],
            "description": "Пятое место из-за неудачного первого выхода (серия RPI). Рост популярности с 2018 после Delta Home Series. 5 лет гарантии — нехарактерно для бюджетников.",
            "advantages": "5 лет гарантии; обновленная линейка Home Series; улучшенное качество после 2018.",
            "features": "Бюджетные инверторы с расширенной гарантией.",
            "application": "Домашние солнечные системы; бюджетный сегмент.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/delta_pv_invertor.webp",
        },
        {
            "category_class": "Бюджетный класс",
            "manufacturer": "Ingeteam",
            "country": "Испания",
            "models": ["1Play TL-M 2,5-6 кВт"],
            "description": "Единственный европеец в бюджетном списке. Отменные характеристики, но проигрывает китайцам по цене. Все модели отличаются европейским качеством, многофункциональностью и информативным мониторингом.",
            "advantages": "Европейское качество; многофункциональность; информативный мониторинг.",
            "features": "Серия 1Play TL-M, мощность 2,5–6 кВт.",
            "application": "Домашние солнечные станции; сетевые системы.",
            "image": "Инверторы для солнечных батарей - наш рейтинг надежности_files/ingeteam.webp",
        },
    ]


def articles_to_products(articles):
    products = []
    for a in articles:
        for model in a["models"]:
            products.append({
                "id": f"article-{len(products)+1}",
                "source": "mywatt.ru — ТОП-12 инверторов",
                "sheet": a["category_class"],
                "category": "Инверторы (обзор)",
                "manufacturer": a["manufacturer"],
                "series": a["category_class"],
                "model": model,
                "article": "",
                "name": f"{a['manufacturer']} {model}",
                "description": a["description"],
                "characteristics": a["features"],
                "advantages": a["advantages"],
                "application": a["application"],
                "country": a["country"],
                "prices": {},
                "images": [a["image"]] if a.get("image") else [],
                "documentation": [],
            })
    return products


def merge_catalog(price_products, web_products, article_products):
    all_items = price_products + web_products + article_products
    by_key = {}
    for p in all_items:
        key = (
            p.get("manufacturer", "").lower(),
            re.sub(r"\s+", " ", p.get("model", "").lower().strip()),
        )
        if key in by_key:
            existing = by_key[key]
            for field in ("description", "characteristics", "advantages", "application", "country", "article"):
                if not existing.get(field) and p.get(field):
                    existing[field] = p[field]
            if not existing.get("prices") and p.get("prices"):
                existing["prices"] = p["prices"]
            elif p.get("prices"):
                for pk, pv in p["prices"].items():
                    if pk not in existing["prices"]:
                        existing["prices"][pk] = pv
            existing["images"] = list(dict.fromkeys((existing.get("images") or []) + (p.get("images") or [])))
            existing["sources"] = list(dict.fromkeys((existing.get("sources") or [existing.get("source")]) + [p.get("source")]))
        else:
            p["sources"] = [p.get("source")]
            by_key[key] = p
    return list(by_key.values())


def main():
    print("Parsing Excel...")
    price_products, excel_stats = parse_excel()
    print(f"  -> {len(price_products)} products from price list")

    print("Parsing bobrovsolar...")
    web_products = parse_bobrovsolar()
    print(f"  -> {len(web_products)} products from website")

    print("Processing inverter article...")
    article_products = articles_to_products(get_inverter_articles())
    print(f"  -> {len(article_products)} entries from article")

    catalog = merge_catalog(price_products, web_products, article_products)

    categories = {}
    for p in catalog:
        cat = p.get("category", "Прочее")
        categories[cat] = categories.get(cat, 0) + 1

    report = {
        "generated_at": datetime.now().isoformat(),
        "sources": [
            {"file": "Прайс_лист_SUNWAYS_29.04.26.xlsx", "type": "Excel прайс", "items": len(price_products), "sheets": excel_stats["sheets"]},
            {"file": "Инверторы для солнечных батарей - наш рейтинг надежности.html", "type": "Статья", "items": len(article_products), "brands": 12},
            {"file": "Аккумуляторы 200 Ач...html (bobrovsolar.ru)", "type": "Сайт", "items": len(web_products)},
            {"file": "каталог панели инвектора.pdf", "type": "PDF каталог", "items": 0, "note": "Требует OCR (китайский текст)"},
            {"file": "пример описания.html", "type": "Эталон описания", "items": 0, "note": "Использован как шаблон структуры карточки"},
            {"file": "ИБП для дома и бизнеса...html", "type": "Сайт ИБП", "items": 0, "note": "Отдельная категория, не включена в текущую выгрузку"},
            {"file": "Top 10 Solar Panel Manufacturers...html", "type": "Статья EN", "items": 0, "note": "Справочно по производителям панелей"},
        ],
        "statistics": {
            "total_products": len(catalog),
            "from_price_list": len(price_products),
            "from_article": len(article_products),
            "from_websites": len(web_products),
            "categories": categories,
            "excel_rows_processed": excel_stats["rows"],
            "excel_sections_skipped": excel_stats["skipped_sections"],
        },
        "price_rules": {
            "retail": ["РРЦ", "РЦ", "Retail", "Цена Розница"],
            "dealer": ["РФЦ", "Dealer", "Цена Дилер"],
            "currencies": {"юань": "внутренняя информация", "USD": "внутренняя информация", "руб": "основная цена для сайта"},
        },
    }

    payload = {"report": report, "products": catalog}
    json_path = OUT_DIR / "catalog.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Saved {json_path} ({len(catalog)} products)")
    return payload


if __name__ == "__main__":
    main()
